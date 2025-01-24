# library_app/management/commands/import_data.py
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from library_app.models import Student, Staff, Book, BookCopy

class Command(BaseCommand):
    help = 'Import or update data from Excel files (books, students, staff), including multiple copies.'

    def add_arguments(self, parser):
        parser.add_argument('--students', type=str, help='Path to students Excel file')
        parser.add_argument('--staff', type=str, help='Path to staff Excel file')
        parser.add_argument('--books', type=str, help='Path to books Excel file')

    def handle(self, *args, **options):
        if options['students']:
            self.import_students(options['students'])
        if options['staff']:
            self.import_staff(options['staff'])
        if options['books']:
            self.import_books(options['books'])

    def import_students(self, path):
        wb = load_workbook(filename=path)
        sheet = wb.active
        # Assuming columns: [Student id, first name, last name, Year Group, Section]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # skip empty rows
                continue
            student_id, first_name, last_name, year_group, section = row
            Student.objects.update_or_create(
                student_id=student_id,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'year_group': year_group,
                    'section': section
                }
            )
        self.stdout.write(self.style.SUCCESS('Successfully imported/updated students.'))

    def import_staff(self, path):
        wb = load_workbook(filename=path)
        sheet = wb.active
        # Assuming columns: [Staff ID, Name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            staff_id, name = row
            Staff.objects.update_or_create(
                staff_id=staff_id,
                defaults={'name': name}
            )
        self.stdout.write(self.style.SUCCESS('Successfully imported/updated staff.'))

    def import_books(self, path):
        wb = load_workbook(filename=path)
        sheet = wb.active
        # Example columns: [Book Name, Image path, ISBN, Rack, Row, Column, Total Copies]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            name, image, isbn, rack, row_, column_, total_copies = row

            # Create or update Book
            book, _ = Book.objects.update_or_create(
                isbn=isbn,
                defaults={
                    'name': name,
                    'image': image,
                    'rack': rack,
                    'row': row_,
                    'column': column_,
                }
            )

            # Adjust copies
            current_copy_count = book.copies.count()
            diff = total_copies - current_copy_count

            if diff > 0:
                # Create new copies
                for i in range(diff):
                    copy_number = f"{isbn}-{current_copy_count + i + 1}"
                    BookCopy.objects.create(book=book, copy_number=copy_number)
            elif diff < 0:
                # Optionally remove extras or handle differently
                # For example, remove last `abs(diff)` copies that are available
                copies_to_remove = book.copies.filter(is_available=True)[:abs(diff)]
                copies_to_remove.delete()

        self.stdout.write(self.style.SUCCESS('Successfully imported/updated books and copies.'))

# commandes to run the script to update the database:
#
# python manage.py import_data --students path/to/students.xlsx
# python manage.py import_data --staff path/to/staff.xlsx
# python manage.py import_data --books path/to/books.xlsx
